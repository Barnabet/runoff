"""Tabular extraction — statement-for-statement port of packages/engine/src/tabular.ts.

File -> detected tables. Pure detection (detect_islands) is separated from IO so
the heuristics are unit-testable on plain grids. scan_tabular (schema, for
classify/drift/UI) and read_tabular (full rows, for ingest) MUST agree on
slugs/columns/order for the same file.

Runtime shapes are plain dicts with camelCase keys:
  DetectedTable  {"slug", "columns", "rowCount", "sample"}
  TabularScan    {"tables": [...], "skipped": [...]}
  WhColumn       {"name", "type"}   type in {"INTEGER", "REAL", "TEXT"}

The TS code parses CSV with papaparse (dynamicTyping) and XLSX with ExcelJS;
here Python's csv module replicates papaparse's dynamic typing and openpyxl
replaces ExcelJS. See the report for the behavioral mappings.
"""

from __future__ import annotations

import csv
import datetime as _dt
import os
import re
from collections.abc import Callable, Iterable
from typing import Any

from openpyxl import load_workbook

BATCH = 10_000
SAMPLE = 10
PREVIEW = 80


def is_tabular(mime: str, name: str) -> bool:
    m = mime.lower()
    if "csv" in m or "spreadsheetml" in m or "ms-excel" in m:
        return True
    return _extname(name).lower() in (".csv", ".xlsx", ".xls")


def _is_csv(mime: str, name: str) -> bool:
    return "csv" in mime.lower() or _extname(name).lower() == ".csv"


def _extname(name: str) -> str:
    return os.path.splitext(name)[1]


def slugify(name: str) -> str:
    # Strip trailing underscores (punctuation tails) but preserve a leading one:
    # a leading underscore is meaningful — it lets a source column collide with
    # the reserved `_period` warehouse column so header_names can bump it.
    s = re.sub(r"_+", "_", re.sub(r"_+$", "", re.sub(r"[^a-z0-9]+", "_", str(name).lower())))
    safe = s or "table"
    return f"t_{safe}" if re.match(r"^\d", safe) else safe


def _is_empty(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


class _ColumnTyper:
    """Per-column type accumulator: all-int -> INTEGER, all-num -> REAL, else TEXT."""

    def __init__(self) -> None:
        self._saw_value: dict[int, bool] = {}
        self._all_int: dict[int, bool] = {}
        self._all_num: dict[int, bool] = {}

    def update(self, row: list[Any]) -> None:
        for i in range(len(row)):
            v = row[i]
            if _is_empty(v):
                continue
            self._saw_value[i] = True
            # A Python bool is an int subclass, but JS booleans are not numbers —
            # exclude bool to match `typeof v === "number"`.
            num = isinstance(v, (int, float)) and not isinstance(v, bool) and _is_finite(v)
            self._all_num[i] = self._all_num.get(i, True) and num
            self._all_int[i] = self._all_int.get(i, True) and num and _is_integer(v)

    def type_at(self, i: int) -> str:
        if not self._saw_value.get(i, False):
            return "TEXT"
        if self._all_int.get(i, False):
            return "INTEGER"
        if self._all_num.get(i, False):
            return "REAL"
        return "TEXT"


def _is_finite(v: Any) -> bool:
    if isinstance(v, float):
        import math

        return math.isfinite(v)
    return True


def _is_integer(v: Any) -> bool:
    if isinstance(v, bool):
        return False
    if isinstance(v, int):
        return True
    if isinstance(v, float):
        return v.is_integer()
    return False


def _header_names(raw: list[Any]) -> list[str]:
    names: list[str] = []
    used: set[str] = set()
    for i, h in enumerate(raw):
        name = f"column_{i + 1}" if _is_empty(h) else slugify(_js_str(h))
        if name == "_period":
            name = "_period_2"
        candidate = name
        n = 2
        while candidate in used:
            candidate = f"{name}_{n}"
            n += 1
        used.add(candidate)
        names.append(candidate)
    return names


def detect_islands(
    grid: list[list[Any]],
    sheet_slug: str,
    used_slugs: set[str] | None = None,
) -> dict[str, list]:
    """Deterministic island detection: blank-row bands x blank-column runs.

    `used_slugs` accumulates every slug already emitted so callers can enforce
    uniqueness across a whole file (all sheets + islands), not just within one
    sheet — pass a shared set to dedupe across sheets via the `_2`/`_3` convention.

    Returns {"tables": [{"slug", "header", "rows"}], "skipped": [str]}.
    """
    if used_slugs is None:
        used_slugs = set()
    tables: list[dict[str, Any]] = []
    skipped: list[str] = []

    def row_at(r: int) -> list[Any]:
        return grid[r] if 0 <= r < len(grid) and grid[r] is not None else []

    def row_has(r: int) -> bool:
        return any(not _is_empty(c) for c in row_at(r))

    bands: list[list[int]] = []
    for r in range(len(grid)):
        if not row_has(r):
            continue
        if bands and bands[-1][1] == r - 1:
            bands[-1][1] = r
        else:
            bands.append([r, r])

    for r0, r1 in bands:
        max_col = 0
        for r in range(r0, r1 + 1):
            max_col = max(max_col, len(row_at(r)))

        def col_has(c: int, r0: int = r0, r1: int = r1) -> bool:
            for r in range(r0, r1 + 1):
                row = row_at(r)
                if c < len(row) and not _is_empty(row[c]):
                    return True
            return False

        runs: list[list[int]] = []
        for c in range(max_col):
            if not col_has(c):
                continue
            if runs and runs[-1][1] == c - 1:
                runs[-1][1] = c
            else:
                runs.append([c, c])

        for c0, c1 in runs:
            width = c1 - c0 + 1
            height = r1 - r0 + 1

            def cells(r: int, c0: int = c0, width: int = width) -> list[Any]:
                row = row_at(r)
                return [row[c0 + i] if c0 + i < len(row) else None for i in range(width)]

            if width < 2 or height < 2:
                text: list[str] = []
                for r in range(r0, r1 + 1):
                    for v in cells(r):
                        if not _is_empty(v):
                            text.append(_js_str(v))
                skipped.append(" ".join(text)[:PREVIEW])
                continue

            header = _header_names(cells(r0))
            rows: list[list[Any]] = []
            for r in range(r0 + 1, r1 + 1):
                row = cells(r)
                if any(not _is_empty(v) for v in row):
                    rows.append(row)
            base = sheet_slug if len(tables) == 0 else f"{sheet_slug}_{len(tables) + 1}"
            slug = base
            n = 2
            while slug in used_slugs:
                slug = f"{base}_{n}"
                n += 1
            used_slugs.add(slug)
            tables.append({"slug": slug, "header": header, "rows": rows})

    return {"tables": tables, "skipped": skipped}


def _table_from_rows(slug: str, header: list[str], rows: list[list[Any]]) -> dict[str, Any]:
    typer = _ColumnTyper()
    for r in rows:
        typer.update(r)
    return {
        "slug": slug,
        "columns": [{"name": name, "type": typer.type_at(i)} for i, name in enumerate(header)],
        "rowCount": len(rows),
        "sample": rows[:SAMPLE],
    }


# --- papaparse dynamicTyping replica ---------------------------------------

_MAX_FLOAT = 2**53
_FLOAT_RE = re.compile(r"^\s*-?(\d+\.?|\.\d+|\d+\.\d+)([eE][-+]?\d+)?\s*$")
# papaparse's streaming ISO_DATE: only full timestamps with T + timezone match;
# a bare YYYY-MM-DD stays a string.
_ISO_DATE_RE = re.compile(
    r"^((\d{4}-[01]\d-[0-3]\dT[0-2]\d:[0-5]\d:[0-5]\d\.\d+([+-][0-2]\d:[0-5]\d|Z))"
    r"|(\d{4}-[01]\d-[0-3]\dT[0-2]\d:[0-5]\d:[0-5]\d([+-][0-2]\d:[0-5]\d|Z))"
    r"|(\d{4}-[01]\d-[0-3]\dT[0-2]\d:[0-5]\d([+-][0-2]\d:[0-5]\d|Z)))$"
)


def _test_float(s: str) -> bool:
    if _FLOAT_RE.match(s):
        v = float(s)
        if -_MAX_FLOAT < v < _MAX_FLOAT:
            return True
    return False


def _js_number(s: str) -> int | float:
    """parseFloat, kept as a JS number: integral values become int so they
    stringify (and integer-detect) like JS `Number`s; fractional stay float."""
    v = float(s)
    return int(v) if v.is_integer() else v


def _parse_dynamic(value: str) -> Any:
    if value in ("true", "TRUE"):
        return True
    if value in ("false", "FALSE"):
        return False
    if _test_float(value):
        return _js_number(value)
    if _ISO_DATE_RE.match(value):
        return _dt.datetime.fromisoformat(value)
    return None if value == "" else value


def _read_csv_rows(path: str) -> Iterable[list[str]]:
    """Raw CSV rows with papaparse's skipEmptyLines (non-greedy): drop only a
    line that is a single empty field (or a wholly blank line)."""
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for row in csv.reader(fh):
            if len(row) == 0 or (len(row) == 1 and row[0] == ""):
                continue
            yield row


# --- CSV -------------------------------------------------------------------


def _stream_csv(
    path: str,
    on_header: Callable[[list[str]], None],
    on_row: Callable[[list[Any]], None],
) -> None:
    """One pass over a CSV; hands each dynamically-typed data row to `on_row`."""
    header: list[str] | None = None
    for raw in _read_csv_rows(path):
        if header is None:
            header = _header_names(list(raw))
            on_header(header)
            continue
        on_row([_parse_dynamic(v) for v in raw])


def csv_grid(path: str) -> list[list[Any]]:
    """Whole-file CSV grid with NO dynamic typing — plan-path callers own typing."""
    return [list(row) for row in _read_csv_rows(path)]


# --- XLSX ------------------------------------------------------------------


def _to_iso_string(dt: _dt.datetime) -> str:
    """Match JS Date.toISOString(): UTC, millisecond precision, trailing Z."""
    if dt.tzinfo is not None:
        dt = dt.astimezone(_dt.UTC).replace(tzinfo=None)
    ms = dt.microsecond // 1000
    date = f"{dt.year:04d}-{dt.month:02d}-{dt.day:02d}"
    time = f"{dt.hour:02d}:{dt.minute:02d}:{dt.second:02d}.{ms:03d}"
    return f"{date}T{time}Z"


def _js_str(v: Any) -> str:
    """String(v) for the value kinds a grid holds (JS booleans stringify lower-case)."""
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, _dt.datetime):
        return _to_iso_string(v)
    return str(v)


def _cell_value(v: Any) -> Any:
    """Port of sourcePack.cellValue for openpyxl (data_only) cell values.

    ExcelJS yields JS Dates and rich-text/formula objects; openpyxl with
    data_only=True already resolves formulas to their cached primitive and
    rich text to plain strings, so only primitives + datetimes reach here.
    """
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        return _to_iso_string(_as_datetime(v))
    if isinstance(v, str):
        return v
    return str(v)


def _as_datetime(v: Any) -> _dt.datetime:
    if isinstance(v, _dt.datetime):
        return v
    if isinstance(v, _dt.date):
        return _dt.datetime(v.year, v.month, v.day)
    if isinstance(v, _dt.time):
        return _dt.datetime(1970, 1, 1, v.hour, v.minute, v.second, v.microsecond)
    return v


def xlsx_grids(path: str) -> list[dict[str, Any]]:
    """Load every sheet's value grid. Returns [{"slug", "grid"}]."""
    out: list[dict[str, Any]] = []
    used_slugs: set[str] = set()
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        raw_name = ws.title
        slug = slugify(raw_name)
        n = 2
        while slug in used_slugs:
            slug = f"{slugify(raw_name)}_{n}"
            n += 1
        used_slugs.add(slug)
        grid: list[list[Any]] = []
        for values in ws.iter_rows(values_only=True):
            grid.append([None if v is None else _cell_value(v) for v in values])
        out.append({"slug": slug, "grid": grid})
    return out


def _detect_xlsx(grids: list[dict[str, Any]]) -> dict[str, list]:
    """Detect all islands across all sheets; per-file slug uniqueness."""
    tables: list[dict[str, Any]] = []
    skipped: list[str] = []
    # One shared set across every sheet so island slugs are unique per file: a
    # sheet-level slug can still collide with an earlier sheet's suffixed island.
    used_slugs: set[str] = set()
    for entry in grids:
        d = detect_islands(entry["grid"], entry["slug"], used_slugs)
        tables.extend(d["tables"])
        skipped.extend(d["skipped"])
    return {"tables": tables, "skipped": skipped}


# --- public API ------------------------------------------------------------


def scan_tabular(path: str, mime: str, name: str) -> dict[str, Any]:
    if _is_csv(mime, name):
        state: dict[str, Any] = {"header": [], "count": 0}
        typer = _ColumnTyper()
        sample: list[list[Any]] = []

        def on_header(h: list[str]) -> None:
            state["header"] = h

        def on_row(row: list[Any]) -> None:
            typer.update(row)
            if len(sample) < SAMPLE:
                sample.append(row)
            state["count"] += 1

        _stream_csv(path, on_header, on_row)
        header = state["header"]
        if not header or state["count"] == 0:
            return {"tables": [], "skipped": []}
        return {
            "tables": [
                {
                    "slug": slugify(re.sub(r"\.[^.]+$", "", name)),
                    "columns": [{"name": n, "type": typer.type_at(i)} for i, n in enumerate(header)],
                    "rowCount": state["count"],
                    "sample": sample,
                }
            ],
            "skipped": [],
        }
    detected = _detect_xlsx(xlsx_grids(path))
    return {
        "tables": [_table_from_rows(t["slug"], t["header"], t["rows"]) for t in detected["tables"]],
        "skipped": detected["skipped"],
    }


def scan_sample(scan: dict[str, Any]) -> str:
    """Compact text sample of a scan for the classifier prompt (<=2,000 chars)."""
    parts: list[str] = []
    for t in scan["tables"]:
        cols = ", ".join(f"{c['name']} ({c['type']})" for c in t["columns"])
        rows = [
            " | ".join("" if v is None else _js_str(v) for v in r)
            for r in t["sample"]
        ]
        parts.append(f"### {t['slug']} — {t['rowCount']} rows\ncolumns: {cols}\n" + "\n".join(rows))
    if scan["skipped"]:
        parts.append(f"(skipped {len(scan['skipped'])} text fragment(s))")
    return "\n\n".join(parts)[:2000]


def read_tabular(
    path: str,
    mime: str,
    name: str,
    on_table: Callable[[dict[str, Any]], Callable[[list[list[Any]]], None]],
) -> None:
    if _is_csv(mime, name):
        # Scan pass first (types/count come from the whole file), then stream batches.
        scan = scan_tabular(path, mime, name)
        if not scan["tables"]:
            return
        emit = on_table(scan["tables"][0])
        batch: list[list[Any]] = []

        def on_row(row: list[Any]) -> None:
            nonlocal batch
            batch.append([None if isinstance(v, str) and v.strip() == "" else v for v in row])
            if len(batch) >= BATCH:
                emit(batch)
                batch = []

        _stream_csv(path, lambda _h: None, on_row)
        if batch:
            emit(batch)
        return
    detected = _detect_xlsx(xlsx_grids(path))
    for t in detected["tables"]:
        emit = on_table(_table_from_rows(t["slug"], t["header"], t["rows"]))
        for i in range(0, len(t["rows"]), BATCH):
            emit(t["rows"][i : i + BATCH])
